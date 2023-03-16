import openpyxl

# Указываем путь к файлу Excel
path = 'E:\\Altayka\\altayka.xlsx'

# Открываем файл Excel
subcategory_dict = {}
workbook = openpyxl.load_workbook(path)


# Выбираем лист, на котором будем производить поиск
worksheet = workbook['1']

# Создаем новый лист для результатов поиска
results_sheet = workbook.create_sheet("3")

# Задаем значение, которое нужно найти
search_value = input('Введите тег для поиска:')

# Создаем список для хранения найденных ячеек
found_cells = []

# Проходимся по всем строкам и столбцам и ищем нужное значение
# for row in worksheet.iter_rows():
#     for cell in row:
#         # if cell.value and search_value in str(cell.value): # Поиск по значению строки
#         if cell.value and cell.value == search_value: # точный поиск
#             found_cells.append(cell)

# Проходимся по всем строкам и столбцам и ищем нужное значение
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value and cell.value == search_value:
            found_cells.append(row)
            break  # останавливаем проход по строке, если найдено нужное значение

# Выводим найденные строки на экран
for row in found_cells:
    for cell in row:
        print(cell.value, end='\t')
    print()

# # Выводим найденные ячейки на экран
# for cell in found_cells:
#     print(cell.value)

# Сохраняем результаты в файл
workbook.save('altayka.xlsx')
