import openpyxl
import pprint as pp
import time
from appium import webdriver
from selenium import webdriver
from typing import Optional
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.firefox.options import Options as Op
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as E
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager

# открываем файл Excel
filename = 'altayka.xlsm'
subcategory_dict = {}
workbook = openpyxl.load_workbook(filename, data_only=True)

# выбираем нужный лист
sheet = workbook['1']

# Чтение содержимого ячеек по заданным диапазонам(строка, столбец)
max_rows = sheet.max_row
for i in range(14, max_rows + 1):
    ordinal_number = sheet.cell(row=i, column=1).value
    sku = sheet.cell(row=i, column=25).value
    subcategory = sheet.cell(row=i, column=4).value
    # Пропуск печати пустых ячеек
    if not sku:
        continue
    # Нумерация
    if not ordinal_number:
        break
    # Добавление к одинаковым ключам, разные значения (в столбец)
    if subcategory not in subcategory_dict:
        subcategory_dict[subcategory] = [sku]
    else:
        subcategory_dict[subcategory].append(sku)

    # Печать в столбик (ключ + все его значения)


# pp.pprint(subcategory_dict)
# print(subcategory_dict)

def summ_numbers_key():
    "Суммирование значений одинаковых ключей"
    my_dict = subcategory_dict
    my_dict['восп'] = sum(my_dict['восп'])
    my_dict['оп'] = sum(my_dict['оп'])
    my_dict['увп'] = sum(my_dict['увп'])
    my_dict['завед'] = sum(my_dict['завед'])
    my_dict['пр пп'] = sum(my_dict['пр пп'])
    print(my_dict)


count1 = count2 = count3 = count4 = count5 = count6 = count7 = count8 = count9 = count10 = \
    count11 = count12 = count13 = count14 = count15 = count16 = 0

for values in subcategory_dict.values():
    for value in values:
        if value < 16000:
            count1 += value
            count1 = round(count1, 2)
        elif 16242.1 < value < 18680.0:
            count2 += value
        elif 18680.1 < value < 19490.0:
            count3 += value
        elif 19490.1 < value < 20300.0:
            count4 += value
        elif 20300.1 < value < 21110.0:
            count5 += value
        elif 21110.1 < value < 23550.0:
            count6 += value
        elif 23550.1 < value < 24360.0:
            count7 += value
        elif 24360.1 < value < 25990.0:
            count8 += value
        elif 25990.1 < value < 27610.0:
            count9 += value
        elif 27610.1 < value < 29240.0:
            count10 += value
        elif 29240.1 < value < 30860.0:
            count11 += value
        elif 30860.1 < value < 32480.0:
            count12 += value
        elif 32480.1 < value < 34110.0:
            count13 += value
        elif 34110.1 < value < 35730.0:
            count14 += value
        elif 35730.1 < value < 37360.0:
            count15 += value
        elif 37360.1 < value < 38980.0:
            count16 += value
counts = [count1, count2, count3, count4, count5, count6, count7, count8, count9, count10,
          count11, count12, count13, count14, count15, count16]
# Вывод всех переменных
# for i, count in enumerate(counts):
#     print(f"count{i+1}: {round(count, 2)}")
# Вывод одной переменной
# print(round(count1, 2))

# Вход в DOXELL
exec_path = r"/Altayka-Doxell/geckodriver.exe"
URL = "https://obr.doxcell.ru:8182/analyzzp/index.jsp"
wait_time_out = 15
driver_manager = GeckoDriverManager()
driver_manager.install()
driver = webdriver.Firefox(service=driver_manager.service)
driver.get(URL)

# Ищем поле ввода логина
input_field = W(driver, wait_time_out).until(E.presence_of_element_located((By.NAME, "j_username")))
input_field.send_keys("gnovokuz")

# Ищем поле ввода пароля
input_field = W(driver, wait_time_out).until(E.presence_of_element_located((By.NAME, "j_password")))
input_field.send_keys("WG7RM3")

# Нажимаем клавишу Enter, чтобы войти в систему
input_field.send_keys(Keys.ENTER)
time.sleep(3)
print('Ok')

element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div[5]/div[4]/div[1]/table/tbody/tr[1]/td[1]/img[6]")))
element.click()
time.sleep(3)
# element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div[5]/div[4]/div[3]/table[2]/tbody/tr[1]/td[2]/select/option[48]")))
# element.click()
# time.sleep(3)
# element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div[5]/div[4]/div[3]/div/table/tbody/tr/td[1]")))
# element.click()
# time.sleep(3)
# element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div[5]/div[4]/div[4]/table[2]/tbody/tr[2]/td[2]/select/option[10]")))
# element.click()
# time.sleep(3)
# element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div[5]/div[4]/div[4]/div/table/tbody/tr[11]/td[1]")))
# element.click()
# time.sleep(3)
# element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, '//*[@id="ref_8440"]')))
# element.click()

# def getOverlappingElement(driver: W, element: WebElement) -> Optional[W]:
#     rect = element.rect
#     result = driver.execute_script("return document.elementFromPoint(arguments[0], arguments[1]);",
#                                    rect['x'] + rect['width'] // 2, rect['y'] + rect['height'] // 2)
#     if result == element:
#         result = None
#     return result
#
#
# overlapping_element = getOverlappingElement(W(driver, wait_time_out), element)
# if overlapping_element:
#     print("Найден перекрывающий элемент:", overlapping_element)
# else:
#     print("Перекрывающих элементов не найдено")



# Ввод значений на лист 7
# input_field = W(driver, 10).until(E.presence_of_element_located((By.XPATH, '//*[@id="sform"]/form/table/tbody/tr[1]/td[2]/input')))
# input_field.clear()
# input_field.send_keys(count1)
# input_field = W(driver, 10).until(E.presence_of_element_located((By.XPATH, "")))
# input_field.clear()
# input_field.send_keys("220")
# wait_variable.until(E.element_to_be_clickable((By.XPATH, ""))).click()

# https://obr.doxcell.ru:8180/web/index.jsp
