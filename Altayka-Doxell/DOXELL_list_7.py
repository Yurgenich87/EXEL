import openpyxl
import pprint as pp
import time
from appium import webdriver
from selenium import webdriver
from typing import Optional
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.firefox.options import Options as Op
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as E
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager

# открываем файл Excel
filename = 'altayka.xlsx'
subcategory_dict = {}
workbook = openpyxl.load_workbook(filename, keep_vba=True)

# выбираем нужный лист
sheet = workbook['1']

# Чтение содержимого ячеек по заданным диапазонам(строка, столбец)
max_rows = sheet.max_row
count_ordinal_number = 0
for i in range(14, max_rows + 1):
    ordinal_number = sheet.cell(row=i, column=1).value
    sku = sheet.cell(row=i, column=25).value
    subcategory = sheet.cell(row=i, column=4).value
    # Пропуск печати пустых ячеек
    if not ordinal_number:
        continue
    # Нумерация
    if not ordinal_number:
        break
    # Добавление к одинаковым ключам, разные значения (в столбец)
    if subcategory not in subcategory_dict:
        subcategory_dict[subcategory] = [sku]
    else:
        subcategory_dict[subcategory].append(sku)
    count_ordinal_number += 1
print(f"Количество сотрудников: {count_ordinal_number}")

    # Печать в столбик (ключ + все его значения)
# pp.pprint(subcategory_dict)

count1 = count2 = count3 = count4 = count5 = count6 = count7 = count8 = count9 = count10 = \
    count11 = count12 = count13 = count14 = count15 = count16 = count17 = count18 = count19 = count20 = \
    count21 = count22 = count23 = count24 = count25 = count26 = count27 = count28 = count29 = count30 = \
    count31 = 0


for values in subcategory_dict.values():
    for value in values:
        if value is not None:
            if value < 16242.0:
                count1 += value
            elif 16242.1 < value < 18680.0:
                count2 += value
            elif 18680.1 < value < 19490.0:
                count3 += value
            elif 19490.1 < value < 20300.0:
                count4 += value
            elif 20300.1 < value < 21110.0:
                count5 += value
            elif 21110.1 < value < 23550.0:
                count6 += value
            elif 23550.1 < value < 24360.0:
                count7 += value
            elif 24360.1 < value < 25990.0:
                count8 += value
            elif 25990.1 < value < 27610.0:
                count9 += value
            elif 27610.1 < value < 29240.0:
                count10 += value
            elif 29240.1 < value < 30860.0:
                count11 += value
            elif 30860.1 < value < 32480.0:
                count12 += value
            elif 32480.1 < value < 34110.0:
                count13 += value
            elif 34110.1 < value < 35730.0:
                count14 += value
            elif 35730.1 < value < 37360.0:
                count15 += value
            elif 37360.1 < value < 38980.0:
                count16 += value
            elif 38980.1 < value < 40600.0:
                count17 += value
            elif 40600.1 < value < 42200.0:
                count18 += value
            elif 42200.1 < value < 43900.0:
                count19 += value
            elif 43900.1 < value < 45500.0:
                count20 += value
            elif 45500.1 < value < 47100.0:
                count21 += value
            elif 47100.1 < value < 48700.0:
                count22 += value
            elif 48700.1 < value < 55000.0:
                count23 += value
            elif 55000.1 < value < 75000.0:
                count24 += value
            elif 75000.1 < value < 100000.0:
                count25 += value
            elif 100000.1 < value < 200000.0:
                count26 += value
            elif 200000.1 < value < 400000.0:
                count27 += value
            elif 400000.1 < value < 1000000.0:
                count28 += value
            elif 1000000.1 < value < 2000000.0:
                count29 += value
            elif 2000000.1 < value < 3000000.0:
                count30 += value
            elif 3000000.0 < value:
                count31 += value

counts = [count1, count2, count3, count4, count5, count6, count7, count8, count9, count10, count11, count12,
          count13, count14, count15, count16, count17, count18, count19, count20, count21, count22, count23,
          count24, count25, count26, count27, count28, count29, count30, count31]
# rounded_counts = [round(count) for count in counts]

# Загрузить файл Excel
# workbook = openpyxl.load_workbook('altayka.xlsx')

# # Проверить, существует ли лист "2"
# if '2' in workbook.sheetnames:
#     # Выбрать лист "2"
#     sheet = workbook['2']
# else:
#     # Создать лист "2", если он не существует
#     sheet = workbook.create_sheet('2')
#
# # Заполнить столбец D5:D35 значениями переменных
# counts = [count1, count2, count3, count4, count5, count6, count7, count8, count9, count10, count11, count12,
#           count13, count14, count15, count16, count17, count18, count19, count20, count21, count22, count23,
#           count24, count25, count26, count27, count28, count29, count30, count31]
#
# for i in range(len(counts)):
#     sheet.cell(row=i+5, column=4).value = counts[i]
#
# # Сохранить файл Excel
# workbook.save('altayka.xlsx')
# # Вывод одной переменной
print()

# Вход в DOXELL
exec_path = r"/Altayka-Doxell/geckodriver.exe"
URL = "https://obr.doxcell.ru:8182/analyzzp/monitoring.jsp"
wait_time_out = 10
driver_manager = GeckoDriverManager()
driver_manager.install()
driver = webdriver.Chrome(service=driver_manager.service)
driver.get(URL)
# Ищем поле ввода логина
input_field = W(driver, wait_time_out).until(E.presence_of_element_located((By.NAME, "j_username")))
input_field.send_keys("gnovokuz")

# Ищем поле ввода пароля
input_field = W(driver, wait_time_out).until(E.presence_of_element_located((By.NAME, "j_password")))
input_field.send_keys("WG7RM3")

# Нажимаем клавишу Enter, чтобы войти в систему
input_field.send_keys(Keys.ENTER)

time.sleep(10)
print('Ok')
# Нажатие на кнопку Мониторинг в разрезе сводных отчетов
element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, '//*[@id="tblCollations"]/tbody/tr[1]/td[1]/img[6]')))
element.click()
time.sleep(10)
# Выбор период Январь 2023
element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.CSS_SELECTOR, '.monComp > option:nth-child(49)')))
element.click()
time.sleep(3)
# Выбор предприятий
element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.CSS_SELECTOR, "#tblDataByCompRep > tbody:nth-child(3) > tr:nth-child(1)")))
element.click()
time.sleep(3)
# Собгород казенные
element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.CSS_SELECTOR, "#idSourceGroupByCoRep > option[value='1303']")))
element.click()
time.sleep(3)
# Выбор учреждения
element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.XPATH, "//*[@id='87056']/td[1]")))
element.click()
print('Зашли на выбор листов')
time.sleep(35)
print('time.sleep(40)')
# element = W(driver, 15).until(E.element_to_be_clickable((By.CSS_SELECTOR, '#rightContrPanel > ul:nth-child(1) > li:nth-child(5) > a:nth-child(1) > img:nth-child(1)')))
element = W(driver, 15).until(E.presence_of_element_located((By.CSS_SELECTOR, 'a[href="#"][class="sheetSelector"][style="text-decoration: none; color: grey; font: 14pt Verdana;"]')))
element.click()
print('click() - Ok')


# # Ввод значения
# element = W(driver, wait_time_out).until(E.element_to_be_clickable((By.CSS_SELECTOR, 'td.browsableContentColumn:nth-child(4) > div:nth-child(1) > input:nth-child(1)')))
# element.
# print('Ok')

input("Нажмите интер для завершения")

